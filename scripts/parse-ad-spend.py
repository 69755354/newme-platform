#!/usr/bin/env python3
"""
Parse the Meta Ads campaign Excel from COS and output JSON rows for ad_spend insertion.

Usage:
  python3 scripts/parse-ad-spend.py [path_to_excel]
  
If no path given, downloads from COS using env vars COS_SECRET_ID, COS_SECRET_KEY, COS_BUCKET.
The Excel is expected at: 市场投流/Mahmoud-Na-il-Campaigns-Dec-1-2025-May-31-2026.xlsx
"""
import sys, json, os, tempfile, subprocess
from datetime import datetime, date

try:
    import openpyxl
except ImportError:
    print(json.dumps({"error": "openpyxl is required: pip install openpyxl"}), file=sys.stderr)
    sys.exit(1)

# Column mapping (0-indexed) — adjust based on actual Excel layout
# Expected columns: Campaign Name, Ad Set Name, Ad Name, Date, Amount, Impressions, Clicks
# We auto-detect by header row.

COS_KEY = "市场投流/Mahmoud-Na-il-Campaigns-Dec-1-2025-May-31-2026.xlsx"

def download_from_cos():
    """Download Excel from COS to a temp file and return its path."""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    download_script = os.path.join(script_dir, "cos-download.py")
    
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp_path = tmp.name
    
    result = subprocess.run(
        ["python3", download_script, COS_KEY],
        capture_output=True, timeout=120, env=os.environ
    )
    
    if result.returncode != 0:
        stderr = result.stderr.decode().strip()
        raise Exception(f"Download failed: {stderr}")
    
    with open(tmp_path, "wb") as f:
        f.write(result.stdout)
    
    return tmp_path

def parse_value(v):
    """Convert cell value to appropriate type."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    if s == "" or s == "-":
        return None
    # Try to detect numeric
    try:
        return float(s.replace(",", "").replace(" ", ""))
    except (ValueError, AttributeError):
        return s

def find_header_row(ws):
    """Find the header row by looking for common campaign-related headers."""
    for row in ws.iter_rows(min_row=1, max_row=20, values_only=False):
        values = [str(c.value).strip().lower() if c.value else "" for c in row]
        row_text = " ".join(values)
        if "campaign" in row_text and ("spend" in row_text or "amount" in row_text or "cost" in row_text):
            return row[0].row  # 1-indexed row number
    return 1  # fallback

def build_column_mapping(headers):
    """Map header names to standardized column keys."""
    mapping = {}
    header_lower = [str(h).strip().lower() if h else "" for h in headers]
    
    for idx, h in enumerate(header_lower):
        if "campaign" in h and ("name" in h or h == "campaign"):
            mapping["campaign_name"] = idx
        elif "ad set" in h or "adset" in h:
            mapping["adset_name"] = idx
        elif h == "ad name" or h == "ad":
            mapping["ad_name"] = idx
        elif h in ("date", "day", "spend date", "delivery date"):
            mapping["spend_date"] = idx
        elif h in ("amount", "spend", "cost", "amount spent", "spend aed"):
            mapping["amount"] = idx
        elif h in ("impressions", "impression", "reach"):
            mapping["impressions"] = idx
        elif h in ("clicks", "click", "link clicks"):
            mapping["clicks"] = idx
        elif "currency" in h:
            mapping["currency"] = idx
    
    return mapping

def parse_excel(filepath):
    """Parse the Excel and return list of row dicts."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    
    header_row_num = find_header_row(ws)
    
    # Read header
    header_row = list(ws.iter_rows(min_row=header_row_num, max_row=header_row_num, values_only=True))[0]
    col_map = build_column_mapping(header_row)
    
    if "campaign_name" not in col_map or "amount" not in col_map:
        # Try harder: dump first 10 rows for debugging
        debug_rows = []
        for r in ws.iter_rows(min_row=1, max_row=10, values_only=True):
            debug_rows.append([str(v)[:50] if v else "" for v in r])
        raise Exception(
            f"Could not find required columns (campaign_name, amount). "
            f"Header row: {list(header_row)}. "
            f"Col map: {col_map}. "
            f"First 10 rows: {json.dumps(debug_rows, ensure_ascii=False)}"
        )
    
    rows = []
    for row in ws.iter_rows(min_row=header_row_num + 1, values_only=True):
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue  # skip empty rows
        
        campaign = parse_value(row[col_map["campaign_name"]])
        if not campaign or str(campaign).strip() == "":
            continue  # skip rows without campaign name
        
        amount = parse_value(row[col_map.get("amount")])
        if amount is None or (isinstance(amount, str) and amount.strip() == ""):
            continue  # skip rows without spend data
        if isinstance(amount, str):
            try:
                amount = float(amount.replace(",", ""))
            except ValueError:
                continue
        
        row_data = {
            "campaign_name": str(campaign).strip(),
            "adset_name": str(parse_value(row[col_map.get("adset_name")]) or "").strip() or None,
            "ad_name": str(parse_value(row[col_map.get("ad_name")]) or "").strip() or None,
            "spend_date": str(parse_value(row[col_map.get("spend_date")]) or ""),
            "amount": round(float(amount), 2),
            "currency": str(parse_value(row[col_map.get("currency")]) or "AED").strip(),
            "impressions": int(float(parse_value(row[col_map.get("impressions")]) or 0)),
            "clicks": int(float(parse_value(row[col_map.get("clicks")]) or 0)),
            "source": "meta",
        }
        rows.append(row_data)
    
    wb.close()
    return rows

if __name__ == "__main__":
    try:
        filepath = sys.argv[1] if len(sys.argv) > 1 else download_from_cos()
        rows = parse_excel(filepath)
        print(json.dumps(rows, ensure_ascii=False, default=str))
    except Exception as e:
        print(json.dumps({"error": str(e)}, ensure_ascii=False), file=sys.stderr)
        sys.exit(1)
